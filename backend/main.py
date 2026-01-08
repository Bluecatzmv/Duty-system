from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlmodel import Session, select, create_engine, SQLModel
from sqlalchemy.orm import selectinload
from contextlib import asynccontextmanager
from typing import Annotated, List, Optional
from io import BytesIO 
from datetime import timedelta, date, datetime
import requests
import openpyxl
from collections import defaultdict
from jose import JWTError, jwt
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger

from models import User, Role, UserRead, UserCreate, Schedule, ScheduleRead, DutyType, Staff, Holiday, HolidayType, Redemption, SystemConfig, ImportLog
from auth import verify_password, create_access_token, get_password_hash, ACCESS_TOKEN_EXPIRE_MINUTES, SECRET_KEY, ALGORITHM

# ================= 配置区 =================
sqlite_file_name = "database.db"
sqlite_url = f"sqlite:///{sqlite_file_name}"
connect_args = {"check_same_thread": False}
engine = create_engine(sqlite_url, connect_args=connect_args)

scheduler = AsyncIOScheduler()

# ================= 数据库初始化 =================
def create_db_and_tables():
    SQLModel.metadata.create_all(engine)
    with Session(engine) as session:
        admin = session.exec(select(User).where(User.username == "admin")).first()
        if not admin:
            super_admin = User(username="admin", real_name="超级管理员", password_hash=get_password_hash("admin123"), role=Role.SUPER_ADMIN)
            session.add(super_admin)
            session.commit()

# ================= 辅助函数 =================
def get_session():
    with Session(engine) as session:
        yield session

def core_send_notification():
    with Session(engine) as session:
        # 1. 获取配置
        url_obj = session.get(SystemConfig, "wecom_url")
        tpl_obj = session.get(SystemConfig, "wecom_template")
        
        if not url_obj or not url_obj.value:
            print("【定时任务】未配置 Webhook，跳过发送")
            return {"status": "error", "message": "未配置 Webhook"}
            
        # 2. 获取默认模板（如果没有配置的话）
        template = tpl_obj.value if tpl_obj and tpl_obj.value else "### 今日值班\n> 日期：{date}\n> 总值班：{leader}\n> 技术值班：{tech}\n> 机动值班：{mobile}"
        
        # 3. 获取日期
        today = date.today()
        two_days_ago = today - timedelta(days=2)
        
        # 4. 查询排班人员
        leader_sch = session.exec(select(Schedule).where(Schedule.date == today, Schedule.duty_type == DutyType.LEADER)).first()
        leader_name = leader_sch.staff_name if leader_sch else "无"
        
        tech_sch = session.exec(select(Schedule).where(Schedule.date == today, Schedule.duty_type == DutyType.TECH)).first()
        tech_name = tech_sch.staff_name if tech_sch else "无"
        
        mobile_sch = session.exec(select(Schedule).where(Schedule.date == two_days_ago, Schedule.duty_type == DutyType.TECH)).first()
        mobile_name = mobile_sch.staff_name if mobile_sch else "无"
        
        # 5. 【核心修改】替换模板变量（增加了 {date}）
        try:
            content = template.replace("{date}", str(today)) \
                              .replace("{leader}", leader_name) \
                              .replace("{tech}", tech_name) \
                              .replace("{mobile}", mobile_name) \
                              .replace("{mobile_date}", str(two_days_ago))
        except Exception as e:
            content = f"模版解析错误: {str(e)}"
            
        # 6. 发送请求
        payload = {"msgtype": "markdown", "markdown": {"content": content}}
        try:
            requests.post(url_obj.value, json=payload, timeout=5)
            print(f"【定时任务】通知已发送：{today}")
            return {"status": "success"}
        except Exception as e:
            print(f"【定时任务】发送异常：{e}")
            return {"status": "error", "message": str(e)}

@asynccontextmanager
async def lifespan(app: FastAPI):
    create_db_and_tables()
    scheduler.start()
    with Session(engine) as session:
        time_obj = session.get(SystemConfig, "wecom_time")
        if time_obj and time_obj.value:
            try:
                h, m = map(int, time_obj.value.split(":"))
                scheduler.add_job(core_send_notification, CronTrigger(hour=h, minute=m), id="daily_notify", replace_existing=True)
                print(f"已加载定时任务：每天 {time_obj.value} 发送")
            except: print("定时任务时间格式错误")
    yield
    scheduler.shutdown()

app = FastAPI(lifespan=lifespan)
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

async def get_current_user(token: Annotated[str, Depends(oauth2_scheme)], session: Session = Depends(get_session)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None: raise HTTPException(status_code=401)
    except: raise HTTPException(status_code=401)
    user = session.exec(select(User).where(User.username == username)).first()
    if not user: raise HTTPException(status_code=401)
    return user

async def check_super_admin(user: User = Depends(get_current_user)):
    if user.role != Role.SUPER_ADMIN: raise HTTPException(status_code=403, detail="需要超级管理员权限")
    return user

# ================= 认证 =================
@app.post("/token")
async def login(form_data: Annotated[OAuth2PasswordRequestForm, Depends()], session: Session = Depends(get_session)):
    user = session.exec(select(User).where(User.username == form_data.username)).first()
    if not user or not verify_password(form_data.password, user.password_hash): raise HTTPException(status_code=400, detail="账号或密码错误")
    token = create_access_token(data={"sub": user.username, "role": user.role}, expires_delta=timedelta(minutes=480))
    return {"access_token": token, "token_type": "bearer", "role": user.role}

# ================= 1. 进阶统计接口 =================
@app.get("/stats/advanced")
async def get_advanced_stats(year: int, session: Session = Depends(get_session)):
    start_date = date(year, 1, 1)
    end_date = date(year + 1, 1, 1)
    
    # 1. 获取所有节假日定义
    holidays = session.exec(select(Holiday).where(Holiday.date >= start_date, Holiday.date < end_date)).all()
    # 注意：这里我们只把明确标记为“休”的日子当作不排班的节假日
    holiday_set = {h.date for h in holidays if h.type == HolidayType.HOLIDAY}
    
    # 2. 获取所有排班
    schedules = session.exec(
        select(Schedule).where(Schedule.date >= start_date, Schedule.date < end_date).options(selectinload(Schedule.staff))
    ).all()
    
    # --- 模块一：周番矩阵 (排除节假日 & 按天去重) ---
    weekday_matrix = {} 
    processed_duties = set()
    
    for s in schedules:
        if not s.staff or s.staff.department != "技术中心": continue
        if s.date in holiday_set: continue
        unique_key = (s.staff_name, s.date)
        if unique_key in processed_duties: continue
        processed_duties.add(unique_key)
        
        name = s.staff_name
        if name not in weekday_matrix: weekday_matrix[name] = [0] * 7 
        weekday_idx = s.date.weekday()
        weekday_matrix[name][weekday_idx] += 1
        
    matrix_list = []
    for name, counts in weekday_matrix.items():
        matrix_list.append({"name": name, "counts": counts, "total": sum(counts)})
    matrix_list.sort(key=lambda x: x["total"], reverse=True)
    
    # --- 模块二：节假日 C 位分析 ---
    sorted_holidays = sorted(list(holiday_set))
    holiday_groups = []
    if sorted_holidays:
        current_group = [sorted_holidays[0]]
        for i in range(1, len(sorted_holidays)):
            if sorted_holidays[i] - sorted_holidays[i-1] == timedelta(days=1):
                current_group.append(sorted_holidays[i])
            else:
                holiday_groups.append(current_group)
                current_group = [sorted_holidays[i]]
        holiday_groups.append(current_group)
    
    holiday_analysis = []
    total_holiday_days = len(holiday_set)
    holiday_duty_unique_set = set()
    schedule_map = defaultdict(list)
    
    for s in schedules:
        if s.staff and s.staff.department == "技术中心":
            if s.staff_name not in schedule_map[s.date]:
                schedule_map[s.date].append(s.staff_name)
            if s.date in holiday_set:
                holiday_duty_unique_set.add((s.date, s.staff_name))

    total_holiday_duty_count = len(holiday_duty_unique_set)

    for group in holiday_groups:
        length = len(group)
        if length == 0: continue
        center_dates = set()
        if length % 2 != 0:
            center_idx = length // 2
            center_dates.add(group[center_idx])
        else:
            center_idx1 = length // 2 - 1
            center_idx2 = length // 2
            center_dates.add(group[center_idx1])
            center_dates.add(group[center_idx2])
            
        group_details = []
        for d in group:
            duty_names = schedule_map.get(d, [])
            is_center = d in center_dates
            group_details.append({
                "date": d,
                "names": duty_names,
                "is_center": is_center
            })
            
        start_str = group[0].strftime("%m.%d")
        end_str = group[-1].strftime("%m.%d")
        holiday_analysis.append({
            "name": f"{start_str}-{end_str} 假期",
            "length": length,
            "days": group_details
        })
        
    return {
        "weekday_stats": matrix_list,
        "holiday_stats": {
            "total_days": total_holiday_days,
            "total_duties": total_holiday_duty_count,
            "groups": holiday_analysis
        }
    }

# ================= 基础配置接口 =================
class WeComConfigRequest(SQLModel):
    webhook_url: str; message_template: str; daily_time: Optional[str] = None
@app.get("/config/wecom")
async def get_wecom_config(user: User = Depends(get_current_user), session: Session = Depends(get_session)):
    if user.role not in [Role.ADMIN, Role.SUPER_ADMIN]: raise HTTPException(status_code=403)
    url_obj = session.get(SystemConfig, "wecom_url"); tpl_obj = session.get(SystemConfig, "wecom_template"); time_obj = session.get(SystemConfig, "wecom_time")
    return {"webhook_url": url_obj.value if url_obj else "", "message_template": tpl_obj.value if tpl_obj else "", "daily_time": time_obj.value if time_obj else None}
@app.post("/config/wecom")
async def save_wecom_config(req: WeComConfigRequest, user: User = Depends(get_current_user), session: Session = Depends(get_session)):
    if user.role not in [Role.ADMIN, Role.SUPER_ADMIN]: raise HTTPException(status_code=403)
    url_obj = session.get(SystemConfig, "wecom_url"); 
    if not url_obj: url_obj = SystemConfig(key="wecom_url", value="")
    url_obj.value = req.webhook_url; session.add(url_obj)
    tpl_obj = session.get(SystemConfig, "wecom_template")
    if not tpl_obj: tpl_obj = SystemConfig(key="wecom_template", value="")
    tpl_obj.value = req.message_template; session.add(tpl_obj)
    time_obj = session.get(SystemConfig, "wecom_time")
    if not time_obj: time_obj = SystemConfig(key="wecom_time", value="")
    if req.daily_time:
        time_obj.value = req.daily_time; session.add(time_obj)
        try: h, m = map(int, req.daily_time.split(":")); scheduler.add_job(core_send_notification, CronTrigger(hour=h, minute=m), id="daily_notify", replace_existing=True)
        except: raise HTTPException(status_code=400, detail="时间格式错误")
    else:
        time_obj.value = ""; session.add(time_obj)
        if scheduler.get_job("daily_notify"): scheduler.remove_job("daily_notify")
    session.commit(); return {"message": "配置已保存，定时任务已更新"}
@app.post("/notify/send")
async def send_daily_notification_manual(user: User = Depends(get_current_user)):
    if user.role not in [Role.ADMIN, Role.SUPER_ADMIN]: raise HTTPException(status_code=403)
    return core_send_notification()

# ================= 调休逻辑 =================
@app.get("/compensatory/overview")
async def get_compensatory_overview(year: int, session: Session = Depends(get_session)):
    start_date = date(year, 1, 1); end_date = date(year + 1, 1, 1)
    schedules = session.exec(select(Schedule).where(Schedule.date >= start_date, Schedule.date < end_date).options(selectinload(Schedule.staff))).all()
    stats = {}
    for s in schedules:
        if not s.staff or s.staff.department != "技术中心": continue
        name = s.staff_name
        if name not in stats: stats[name] = {"name": name, "total_earned_days": 0, "total_redeemed_days": 0, "balance": 0, "processed_dates": set()}
        if s.date not in stats[name]["processed_dates"]: stats[name]["total_earned_days"] += 2; stats[name]["processed_dates"].add(s.date)
        stats[name]["total_redeemed_days"] += s.redeemed_count
    result = []
    for name, data in stats.items(): data["balance"] = data["total_earned_days"] - data["total_redeemed_days"]; del data["processed_dates"]; result.append(data)
    return sorted(result, key=lambda x: x["balance"], reverse=True)
@app.get("/compensatory/calendar/{staff_name}")
async def get_compensatory_calendar(staff_name: str, year: int, session: Session = Depends(get_session)):
    start_date = date(year, 1, 1); end_date = date(year + 1, 1, 1)
    schedules = session.exec(select(Schedule).where(Schedule.date >= start_date, Schedule.date < end_date, Schedule.staff_name == staff_name).options(selectinload(Schedule.redemptions))).all()
    date_groups = defaultdict(list)
    for s in schedules: date_groups[s.date].append(s)
    events = []; quota_list = []
    for day_date in sorted(date_groups.keys()):
        sibs = date_groups[day_date]
        total_redeemed_on_day = sum(s.redeemed_count for s in sibs)
        first_s = sibs[0]
        events.append({"title": "值班 (+2)", "start": day_date, "color": "#2080f0", "type": "duty", "schedule_id": first_s.id})
        remaining = 2 - total_redeemed_on_day
        if remaining > 0: quota_list.append({"id": first_s.id, "date": day_date, "remaining": remaining})
        for s in sibs:
            for r in s.redemptions:
                events.append({"title": "已休", "start": r.redeem_date, "color": "#18a058", "type": "leave", "redemption_id": r.id, "from_date": day_date})
    return {"events": events, "quota_list": quota_list}
class RedeemRequest(SQLModel):
    staff_name: str; redeem_date: date; schedule_id: int
@app.post("/compensatory/redeem")
async def redeem_schedule(req: RedeemRequest, session: Session = Depends(get_session)):
    sch = session.get(Schedule, req.schedule_id)
    if not sch: raise HTTPException(status_code=404, detail="额度来源不存在")
    siblings = session.exec(select(Schedule).where(Schedule.staff_name == sch.staff_name, Schedule.date == sch.date)).all()
    total_used = sum(s.redeemed_count for s in siblings)
    if total_used >= 2: raise HTTPException(status_code=400, detail="该值班日期的调休额度(2天)已全部用完")
    redemption = Redemption(schedule_id=sch.id, redeem_date=req.redeem_date)
    sch.redeemed_count += 1
    session.add(redemption); session.add(sch); session.commit()
    return {"message": "登记成功"}
@app.delete("/compensatory/redeem/{redemption_id}")
async def cancel_redeem(redemption_id: int, session: Session = Depends(get_session)):
    r = session.get(Redemption, redemption_id)
    if not r: raise HTTPException(status_code=404)
    sch = session.get(Schedule, r.schedule_id)
    if sch: sch.redeemed_count -= 1; session.add(sch)
    session.delete(r); session.commit()
    return {"message": "已撤销"}

# ================= 人员管理 =================
@app.get("/users/", response_model=List[UserRead])
async def get_all_users(user: User = Depends(check_super_admin), session: Session = Depends(get_session)):
    return session.exec(select(User)).all()
@app.post("/users/", response_model=UserRead)
async def create_user(new_user: UserCreate, user: User = Depends(check_super_admin), session: Session = Depends(get_session)):
    if session.exec(select(User).where(User.username == new_user.username)).first(): raise HTTPException(status_code=400)
    db_user = User(username=new_user.username, real_name=new_user.real_name, password_hash=get_password_hash(new_user.password), role=new_user.role)
    session.add(db_user); session.commit(); session.refresh(db_user); return db_user
@app.delete("/users/{user_id}")
async def delete_user(user_id: int, user: User = Depends(check_super_admin), session: Session = Depends(get_session)):
    target = session.get(User, user_id); 
    if target and target.username != "admin": session.delete(target); session.commit()
    return {"message": "ok"}

# ================= 通讯录 (新增公开接口) =================
@app.get("/contacts/public")
async def get_public_contacts(session: Session = Depends(get_session)):
    # 直接查询 Staff 表，不进行权限校验，允许未登录访问
    # 按部门排序，方便查看
    return session.exec(select(Staff).order_by(Staff.department, Staff.real_name)).all()

# ================= 通讯录管理 (新增接口) =================

class ContactUpdateRequest(SQLModel):
    real_name: str
    phone: Optional[str] = None
    department: Optional[str] = None

@app.put("/contacts/{staff_id}")
async def update_contact(
    staff_id: int, 
    req: ContactUpdateRequest, 
    user: User = Depends(get_current_user), # 必须登录
    session: Session = Depends(get_session)
):
    # 权限校验：只有管理员或超级管理员能改
    if user.role not in [Role.ADMIN, Role.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="权限不足")
        
    staff = session.get(Staff, staff_id)
    if not staff:
        raise HTTPException(status_code=404, detail="人员不存在")
    
    # 更新数据
    staff.real_name = req.real_name
    staff.phone = req.phone
    staff.department = req.department
    
    session.add(staff)
    session.commit()
    return {"msg": "更新成功"}

@app.delete("/contacts/{staff_id}")
async def delete_contact(
    staff_id: int,
    user: User = Depends(get_current_user),
    session: Session = Depends(get_session)
):
    if user.role not in [Role.ADMIN, Role.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="权限不足")
        
    staff = session.get(Staff, staff_id)
    if not staff:
        raise HTTPException(status_code=404)
        
    # 注意：如果该员工有排班记录，通常不建议物理删除，或者数据库会报错(外键约束)。
    # 这里简单处理：如果有排班，则不允许删除，或者只删除Staff表记录(取决于你的数据库级联设置)
    # 为了安全，这里我们只尝试删除 Staff
    try:
        session.delete(staff)
        session.commit()
    except Exception as e:
        raise HTTPException(status_code=400, detail="无法删除：该人员可能存在关联的排班记录，请先清理排班。")
        
    return {"msg": "删除成功"}

# ================= 节假日管理 (核心修改) =================
@app.get("/holidays/", response_model=List[Holiday])
async def get_holidays(year: int, month: int, session: Session = Depends(get_session)):
    start = date(year, month, 1); end = date(year+1, 1, 1) if month==12 else date(year, month+1, 1)
    return session.exec(select(Holiday).where(Holiday.date >= start, Holiday.date < end)).all()

# 新增：更灵活的批量请求模型
class HolidayBatchRequest(SQLModel):
    start_date: date
    end_date: date
    # 选项1：设置节假日类型 (可选)
    update_type: bool = False # 标记是否更新类型
    name: Optional[str] = None
    type: Optional[HolidayType] = None
    # 选项2：设置保障期 (可选)
    update_guarantee: bool = False # 标记是否更新保障期
    is_guarantee: bool = False
    guarantee_name: Optional[str] = None

@app.post("/holidays/batch")
async def create_holiday_batch(req: HolidayBatchRequest, user: User = Depends(get_current_user), session: Session = Depends(get_session)):
    if req.start_date > req.end_date: 
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期")
    
    curr = req.start_date
    while curr <= req.end_date:
        ext = session.exec(select(Holiday).where(Holiday.date == curr)).first()
        
        is_new = False
        if not ext:
            ext = Holiday(date=curr)
            is_new = True
        
        # 1. 更新休/班信息
        if req.update_type:
            ext.name = req.name
            ext.type = req.type # 如果选了“清除设置”，这里会变成 None
            
        # 2. 更新保障期信息
        if req.update_guarantee:
            ext.is_guarantee = req.is_guarantee
            ext.guarantee_name = req.guarantee_name if req.is_guarantee else None

        # 3. 自动清理“无效”记录 (核心修复)
        # 如果 Type 是空的，并且也不是保障期，说明这条记录已经没有存在的意义了
        if ext.type is None and not ext.is_guarantee:
            if not is_new:
                # 如果是旧记录，从数据库删除
                session.delete(ext)
            # 如果是新记录，则不执行 session.add，直接忽略
        else:
            # 如果还有效（比如只是清除了休假但它是保障期，或者反之），则保存
            session.add(ext)

        curr += timedelta(days=1)
        
    session.commit()
    return {"msg": "ok"}

@app.delete("/holidays/{holiday_id}")
async def delete_holiday(holiday_id: int, user: User = Depends(get_current_user), session: Session = Depends(get_session)):
    t = session.get(Holiday, holiday_id) # 这里也要改
    if t: 
        session.delete(t)
        session.commit()
    return {"msg": "ok"}

# ================= 导入导出 =================
@app.post("/contacts/import")
async def import_contacts(file: UploadFile = File(...), user: User = Depends(get_current_user), session: Session = Depends(get_session)):
    c = await file.read(); wb = openpyxl.load_workbook(BytesIO(c)); sheet = wb.active
    names = set(); u=0; cr=0
    for r in sheet.iter_rows(min_row=2, values_only=True):
        n=r[0]; d=r[1]; p=str(r[2]) if r[2] else ""; 
        if not n: continue
        n=n.strip(); names.add(n)
        s = session.exec(select(Staff).where(Staff.real_name == n)).first()
        if s: s.department=d; s.phone=p; session.add(s); u+=1
        else: session.add(Staff(real_name=n, department=d, phone=p)); cr+=1
    all_s = session.exec(select(Staff)).all()
    cd=0
    for s in all_s:
        if s.real_name not in names: session.delete(s); cd+=1
    session.commit(); return {"msg": f"Add {cr}, Upd {u}, Del {cd}"}

# ================= 排班 CRUD =================
@app.get("/schedules/", response_model=List[ScheduleRead])
async def read_schedules(year: int, month: int, session: Session = Depends(get_session)):
    start = date(year, month, 1); end = date(year+1, 1, 1) if month==12 else date(year, month+1, 1)
    s = session.exec(select(Schedule).where(Schedule.date >= start, Schedule.date < end).options(selectinload(Schedule.staff))).all()
    res = []
    for i in s: res.append(ScheduleRead(id=i.id, date=i.date, duty_type=i.duty_type, staff_name=i.staff_name, staff_phone=i.staff.phone if i.staff else "", redeemed_count=i.redeemed_count))
    return res

@app.put("/schedules/{sid}")
async def update_schedule(sid: int, d: ScheduleRead, u: User = Depends(get_current_user), s: Session = Depends(get_session)):
    db = s.get(Schedule, sid)
    if not db: raise HTTPException(404)
    
    # 1. 处理日期和排班类型
    dt = d.date
    if isinstance(dt, str): dt = datetime.strptime(dt, "%Y-%m-%d").date()
    db.date = dt
    db.duty_type = d.duty_type
    
    # 2. 获取新名字和旧名字
    new_staff_name = d.staff_name.strip()
    old_staff_name = db.staff_name  # 记录修改前的名字
    
    # 3. 查找人员是否存在
    st = s.exec(select(Staff).where(Staff.real_name == new_staff_name)).first()
    
    if not st: 
        # 【情况 A：新人员不存在】 -> 创建新人员
        # 这种情况下，我们信任前端传来的电话（因为数据库里没记录）
        st = Staff(real_name=new_staff_name, phone=d.staff_phone)
        s.add(st)
        s.commit()
        s.refresh(st)
    else: 
        # 【情况 B：人员已存在】 -> 关键修复逻辑
        # 只有当【名字没有发生变化】时，才允许通过排班表单修改电话号码。
        # 如果【名字变了】（说明是换班），由于前端往往残留着上一个人的电话，
        # 所以我们为了保护该人员原有的数据，忽略前端传来的 phone，除非你真的想改，需要先换人保存，再点开改电话。
        if new_staff_name == old_staff_name:
            if d.staff_phone and st.phone != d.staff_phone: 
                st.phone = d.staff_phone
                s.add(st)
        # else: 如果名字变了，什么都不做，直接使用 st 原有的电话，防止被脏数据覆盖
    
    # 4. 更新排班记录关联
    db.staff_name = new_staff_name
    db.staff_id = st.id
    s.add(db)
    s.commit()
    return {"msg": "ok"}

@app.delete("/schedules/{sid}")
async def delete_schedule(sid: int, u: User = Depends(get_current_user), s: Session = Depends(get_session)):
    t = s.get(Schedule, sid); 
    if t: s.delete(t); s.commit()
    return {"msg": "ok"}
@app.get("/imports/history")
async def get_import_history(limit: int = 20, session: Session = Depends(get_session)):
    # 按时间倒序返回最近的导入记录
    return session.exec(select(ImportLog).order_by(ImportLog.import_time.desc()).limit(limit)).all()

# ================= 优化：排班导入 (核心修改) =================
@app.post("/schedules/import_excel")
async def import_excel(
    file: UploadFile = File(...), 
    is_overwrite: bool = True,  # 新增参数：是否覆盖
    user: User = Depends(get_current_user), 
    session: Session = Depends(get_session)
):
    # 1. 读取文件
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    sheet = wb.active
    
    # 2. 预处理数据（先读到内存，不急着写库）
    new_schedules = []
    involved_dates = set()
    
    cm = { 2: DutyType.LEADER, 3: DutyType.TECH, 4: DutyType.DAY, 5: DutyType.NIGHT, 6: DutyType.NIGHT_TRAINEE, 7: DutyType.UPDATE, 8: DutyType.TRAINEE }
    
    for r in sheet.iter_rows(min_row=2, values_only=True):
        dv = r[0]
        if not dv: continue
        # 处理日期格式
        sd = dv if isinstance(dv, (datetime, date)) else datetime.strptime(str(dv), "%Y-%m-%d").date()
        involved_dates.add(sd)
        
        for idx, dt in cm.items():
            if idx >= len(r): break
            nm = r[idx]
            if nm and str(nm).strip():
                nm = str(nm).strip()
                # 自动关联或创建 Staff
                st = session.exec(select(Staff).where(Staff.real_name == nm)).first()
                if not st: 
                    st = Staff(real_name=nm)
                    session.add(st)
                    session.commit()
                    session.refresh(st)
                
                # 准备排班对象
                sch = Schedule(date=sd, staff_name=nm, staff_id=st.id, duty_type=dt)
                new_schedules.append(sch)
    
    if not new_schedules:
        return {"msg": "文件中未读取到有效排班数据"}

    # 3. 执行覆盖逻辑 (核心解决重复问题)
    min_date = min(involved_dates)
    max_date = max(involved_dates)
    
    if is_overwrite:
        # 删除该时间段内数据库中已有的记录，防止叠加
        # 注意：这里会删除范围内所有的排班，实现“整月替换”的效果
        existing = session.exec(
            select(Schedule).where(Schedule.date >= min_date, Schedule.date <= max_date)
        ).all()
        for e in existing:
            session.delete(e)
        session.flush() # 立即执行删除

    # 4. 写入新数据
    for item in new_schedules:
        session.add(item)
    
    # 5. 记录日志
    desc = f"覆盖: {min_date} 至 {max_date}" if is_overwrite else f"追加: {min_date} 至 {max_date}"
    log = ImportLog(
        filename=file.filename,
        import_type="schedule",
        operator_name=user.real_name,
        description=desc,
        row_count=len(new_schedules)
    )
    session.add(log)
    
    session.commit()
    return {"msg": f"成功导入 {len(new_schedules)} 条排班记录", "description": desc}

# ================= 优化：通讯录导入 =================
@app.post("/contacts/import")
async def import_contacts(
    file: UploadFile = File(...), 
    is_overwrite: bool = False, # 通讯录通常是追加/更新
    user: User = Depends(get_current_user), 
    session: Session = Depends(get_session)
):
    c = await file.read(); wb = openpyxl.load_workbook(BytesIO(c)); sheet = wb.active
    names = set(); u=0; cr=0
    
    # 如果选择覆盖模式，逻辑可以更激进（这里暂时保持保留已有人员ID，更新信息）
    # 如果用户真的想完全重置通讯录，建议提供一个“清空通讯录”的按钮，这里为了安全起见，逻辑稍微保守
    
    for r in sheet.iter_rows(min_row=2, values_only=True):
        n=r[0]; d=r[1]; p=str(r[2]) if r[2] else ""; 
        if not n: continue
        n=n.strip(); names.add(n)
        s = session.exec(select(Staff).where(Staff.real_name == n)).first()
        if s: 
            s.department=d; s.phone=p; session.add(s); u+=1
        else: 
            session.add(Staff(real_name=n, department=d, phone=p)); cr+=1
            
    # 如果需要处理“删除Excel中没有的人”，需要谨慎
    # 此处逻辑保持：更新存在的，添加不存在的。
    
    # 记录日志
    log = ImportLog(
        filename=file.filename,
        import_type="contact",
        operator_name=user.real_name,
        description=f"更新通讯录: 新增{cr}人, 更新{u}人",
        row_count=cr + u
    )
    session.add(log)
    
    session.commit()
    return {"msg": f"新增 {cr}, 更新 {u}"}

# ================= 年度统计 =================
@app.get("/stats/yearly")
async def get_yearly_stats(year: int, session: Session = Depends(get_session)):
    start_date = date(year, 1, 1); end_date = date(year + 1, 1, 1)
    holidays_query = session.exec(select(Holiday).where(Holiday.date >= start_date, Holiday.date < end_date)).all()
    # 统计仅关注“休”
    holiday_set = {h.date for h in holidays_query if h.type == HolidayType.HOLIDAY}
    workday_set = {h.date for h in holidays_query if h.type == HolidayType.WORKDAY}
    schedules = session.exec(select(Schedule).where(Schedule.date >= start_date, Schedule.date < end_date).options(selectinload(Schedule.staff))).all()
    stats = {}
    for s in schedules:
        if not s.staff or s.staff.department != "技术中心": continue
        name = s.staff_name
        if name not in stats: stats[name] = {"name": name, "worked_dates": set(), "total": 0, "holiday_count": 0, "weekend_count": 0, "weekday_count": 0, "months": [0] * 12, "details": []}
        is_holiday = s.date in holiday_set
        is_weekend = (s.date.weekday() >= 5) and (s.date not in workday_set)
        if s.date not in stats[name]["worked_dates"]:
            if is_holiday: stats[name]["holiday_count"] += 1
            elif is_weekend: stats[name]["weekend_count"] += 1
            else: stats[name]["weekday_count"] += 1
            stats[name]["months"][s.date.month - 1] += 1
            stats[name]["details"].append({"date": s.date, "type": s.duty_type, "is_holiday": is_holiday, "is_weekend": is_weekend})
        stats[name]["worked_dates"].add(s.date)
    result_list = []
    for name, data in stats.items():
        data["total"] = len(data["worked_dates"]); del data["worked_dates"]; data["details"].sort(key=lambda x: x["date"]); result_list.append(data)
    result_list.sort(key=lambda x: x["total"], reverse=True)
    return result_list
